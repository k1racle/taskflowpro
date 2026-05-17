import json
import sys
from pathlib import Path

try:
    import openpyxl
except Exception as exc:
    print(json.dumps({"error": f"openpyxl import failed: {exc}"}, ensure_ascii=False))
    sys.exit(1)


def normalize(value):
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value)


def resolve_path(arg_list):
    if len(arg_list) >= 2:
        return Path(arg_list[1])

    old_dir = Path('old')
    matches = sorted([p for p in old_dir.iterdir() if p.is_file() and p.suffix.lower() in {'.xlsx', '.xlsm', '.xltx', '.xltm'}])
    if len(matches) == 1:
        return matches[0]
    raise FileNotFoundError('Excel file not found or ambiguous in old/')


def main() -> int:
    try:
        path = resolve_path(sys.argv)
    except Exception as exc:
        print(json.dumps({"error": str(exc)}, ensure_ascii=False))
        return 1

    wb = openpyxl.load_workbook(path, data_only=True)
    result = {"sheets": []}

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        preview = []
        for row in rows[:10]:
            preview.append([normalize(cell) for cell in row])
        result["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "preview": preview,
        })

    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
